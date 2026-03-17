"""
Build institutes table v2:
  1. Combine EXACT matches (1763) + OK review entries (77) as confirmed
  2. Create institutes table with SM display_name + Profile data
  3. Match DB allotment institute_names to institutes via fuzzy matching
  4. Export unknown institutes to Excel
"""
import sqlite3, os, sys, re, math
import pandas as pd
from rapidfuzz import fuzz, process
import unicodedata
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

# ── Paths ─────────────────────────────────────────────────────────────────────
BASE_DIR = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
DB_PATH = os.path.join(BASE_DIR, "data", "neetpg.db")
MATCHED_CSV = "/tmp/matched_institutes_v3.csv"
PROFILES_CSV = "/tmp/profiles_v2.csv"
SM_CSV = "/tmp/seat_matrix.csv"
REVIEW_XLSX = "/sessions/cool-eloquent-davinci/mnt/uploads/College List v5.xlsx"
OUTPUT_DIR = "/sessions/cool-eloquent-davinci/mnt/neetpg-counselling-v2"

# ── Helpers ───────────────────────────────────────────────────────────────────
def norm(s):
    if not isinstance(s, str): return ''
    s = unicodedata.normalize('NFKD', s).encode('ascii', 'ignore').decode()
    return re.sub(r'\s+', ' ', re.sub(r'[^\w\s]', ' ', s.lower())).strip()

GENERIC = {'hospital','medical','college','institute','university','govt',
           'government','district','general','super','speciality','specialty',
           'centre','center','research','national','regional','state','pvt',
           'ltd','private','limited','trust','foundation','society','memorial',
           'of','the','and','in','at'}

def strip_generic(s):
    words = norm(s).split()
    core = [w for w in words if w not in GENERIC and len(w) > 1]
    return ' '.join(core) if core else ' '.join(words)

def to_pin(x):
    if pd.isna(x): return ''
    try: return str(int(float(x)))
    except: return str(x).strip()

def extract_city_from_address(address, institute_name):
    """Extract city name from SM address field."""
    if not isinstance(address, str) or not address.strip():
        return ''
    # Clean the address
    addr = address.strip()
    # Remove the institute name from the beginning if present
    name_lower = institute_name.lower().strip() if isinstance(institute_name, str) else ''
    addr_check = addr.lower().strip()
    if addr_check.startswith(name_lower):
        addr = addr[len(name_lower):].strip().lstrip(',').strip()

    # Split by comma and look for city-like segments
    segs = [s.strip() for s in addr.split(',') if s.strip()]

    # State names to exclude
    STATES = {
        'andhra pradesh','arunachal pradesh','assam','bihar','chhattisgarh','goa',
        'gujarat','haryana','himachal pradesh','jharkhand','karnataka','kerala',
        'madhya pradesh','maharashtra','manipur','meghalaya','mizoram','nagaland',
        'odisha','punjab','rajasthan','sikkim','tamil nadu','telangana','tripura',
        'uttar pradesh','uttarakhand','west bengal','puducherry','delhi',
        'jammu and kashmir','jammu & kashmir','ladakh','chandigarh',
        'andaman and nicobar islands','dadra and nagar haveli','lakshadweep',
        'new delhi',
    }

    # Find last meaningful segment that's not a state or pincode
    for seg in reversed(segs):
        seg_clean = seg.strip()
        seg_lower = seg_clean.lower()
        # Skip pincode
        if re.match(r'^\d{5,6}$', seg_clean):
            continue
        # Skip state names
        seg_no_pin = re.sub(r'[-\s]+\d{4,6}\s*$', '', seg_lower).strip()
        if seg_no_pin in STATES:
            continue
        # Skip very long segments (likely full address, not city)
        if len(seg_clean) > 40:
            continue
        # Skip if it looks like a full address with road/street etc
        if re.search(r'\b(road|street|marg|lane|nagar|colony|sector|phase|block|plot|campus)\b', seg_lower):
            # But could still contain city name at the end
            # Try the last word as potential city
            words = seg_clean.split()
            if words:
                last = words[-1].strip()
                if len(last) > 2 and last.lower() not in STATES and not re.match(r'^\d+$', last):
                    return last
            continue
        # This segment is likely a city/district name
        if len(seg_clean) > 1:
            return seg_clean

    # Fallback: try extracting from the original address
    # Look for common city patterns
    segs_orig = [s.strip() for s in address.split(',') if s.strip()]
    for seg in reversed(segs_orig):
        seg_lower = seg.strip().lower()
        if seg_lower not in STATES and not re.match(r'^\d{5,6}$', seg.strip()) and len(seg.strip()) < 30 and len(seg.strip()) > 1:
            return seg.strip()

    return ''


# ══════════════════════════════════════════════════════════════════════════════
# STEP 1: Build confirmed institutes list
# ══════════════════════════════════════════════════════════════════════════════
print("=" * 70)
print("STEP 1: Building confirmed institutes list")
print("=" * 70)

# Load data
matched = pd.read_csv(MATCHED_CSV)
profiles = pd.read_csv(PROFILES_CSV)
sm = pd.read_csv(SM_CSV)

print(f"Matched CSV: {len(matched)} rows")
print(f"Match status breakdown:")
print(matched['match_status'].value_counts().to_string())

# 1a. EXACT matches
exact_df = matched[matched['match_status'] == 'EXACT'].copy()
print(f"\nEXACT matches: {len(exact_df)}")

# 1b. OK entries from review sheet
wb = openpyxl.load_workbook(REVIEW_XLSX, data_only=True)
review_ws = None
for sname in wb.sheetnames:
    if '要確認' in sname or '候補' in sname:
        review_ws = wb[sname]
        break

ok_inst_codes = set()
if review_ws:
    for row in review_ws.iter_rows(min_row=3, max_col=3):
        judgment = row[0].value
        inst_code = row[2].value
        if judgment and str(judgment).strip().upper() == 'OK' and inst_code:
            ok_inst_codes.add(int(float(inst_code)))

print(f"OK review entries: {len(ok_inst_codes)}")

# Get non-EXACT rows that are OK'd
ok_review_df = matched[
    (matched['match_status'] != 'EXACT') &
    (matched['institute_code'].isin(ok_inst_codes))
].copy()
print(f"OK review rows found in matched CSV: {len(ok_review_df)}")

# For OK review entries, we need to re-match them to the correct profile candidate
# The review sheet used pin-based matching, so for these we re-do the lookup
prof_pin_map = {}
for i, row in profiles.iterrows():
    p = to_pin(row['pincode'])
    if p:
        prof_pin_map.setdefault(p, []).append(i)

def find_best_profile_for_ok(sm_row):
    """For an OK review entry, find the best profile match using pin + name scoring."""
    sm_pin = to_pin(sm_row.get('pincode', ''))
    sm_name = norm(str(sm_row.get('institute_name', '')))
    sm_addr = norm(str(sm_row.get('address', '')))

    cands = []
    if sm_pin and sm_pin in prof_pin_map:
        for idx in prof_pin_map[sm_pin]:
            pr = profiles.iloc[idx]
            pr_name = norm(str(pr.get('college_name', '')))
            pr_addr = norm(str(pr.get('address_profile', '')))
            # Combined score
            sm_combined = sm_name + ' ' + sm_addr
            pr_combined = pr_name + ' ' + pr_addr
            score = fuzz.token_sort_ratio(sm_combined, pr_combined)
            cands.append((score, idx))

    if cands:
        cands.sort(key=lambda x: -x[0])
        return profiles.iloc[cands[0][1]]

    # Fallback: use the profile_name from matched CSV
    prof_name = str(sm_row.get('profile_name', ''))
    if prof_name:
        matches = profiles[profiles['college_name'] == prof_name]
        if len(matches) > 0:
            return matches.iloc[0]

    return None

# Build confirmed list
confirmed = []

# Add EXACT matches
for _, row in exact_df.iterrows():
    inst = {
        'institute_code': int(row['institute_code']),
        'institute_name': str(row['institute_name']),
        'address_sm': str(row['address']),
        'state': str(row['state']),
        'pincode': to_pin(row.get('pincode', '')),
        'match_status': 'EXACT',
    }

    # Extract city from SM address
    inst['city'] = extract_city_from_address(inst['address_sm'], inst['institute_name'])

    # Build display name: SM name + city
    city = inst['city']
    if city and city.lower() not in inst['institute_name'].lower():
        inst['display_name'] = f"{inst['institute_name']}, {city}"
    else:
        inst['display_name'] = inst['institute_name']

    # Profile data (from prof_ prefixed columns)
    for prof_col in ['university', 'fee_yr1', 'fee_yr2', 'fee_yr3',
                     'stipend_yr1', 'stipend_yr2', 'stipend_yr3',
                     'hostel_male', 'hostel_female', 'bond_forfeit',
                     'pwbd_friendly', 'website', 'annual_fee']:
        val = row.get(f'prof_{prof_col}', '')
        if pd.isna(val): val = None
        inst[prof_col] = val

    confirmed.append(inst)

# Add OK review matches
for _, row in ok_review_df.iterrows():
    inst = {
        'institute_code': int(row['institute_code']),
        'institute_name': str(row['institute_name']),
        'address_sm': str(row['address']),
        'state': str(row['state']),
        'pincode': to_pin(row.get('pincode', '')),
        'match_status': 'OK_REVIEW',
    }

    inst['city'] = extract_city_from_address(inst['address_sm'], inst['institute_name'])
    city = inst['city']
    if city and city.lower() not in inst['institute_name'].lower():
        inst['display_name'] = f"{inst['institute_name']}, {city}"
    else:
        inst['display_name'] = inst['institute_name']

    # Find best profile match
    prof_row = find_best_profile_for_ok(row)
    if prof_row is not None:
        for prof_col in ['university', 'fee_yr1', 'fee_yr2', 'fee_yr3',
                         'stipend_yr1', 'stipend_yr2', 'stipend_yr3',
                         'hostel_male', 'hostel_female', 'bond_forfeit',
                         'pwbd_friendly', 'website', 'annual_fee']:
            val = prof_row.get(prof_col, '')
            if pd.isna(val): val = None
            inst[prof_col] = val
    else:
        for prof_col in ['university', 'fee_yr1', 'fee_yr2', 'fee_yr3',
                         'stipend_yr1', 'stipend_yr2', 'stipend_yr3',
                         'hostel_male', 'hostel_female', 'bond_forfeit',
                         'pwbd_friendly', 'website', 'annual_fee']:
            inst[prof_col] = None

    confirmed.append(inst)

# Add remaining unmatched SM entries
confirmed_codes = {c['institute_code'] for c in confirmed}
for _, row in sm.iterrows():
    code = int(row['institute_code'])
    if code in confirmed_codes:
        continue
    inst = {
        'institute_code': code,
        'institute_name': str(row['institute_name']),
        'address_sm': str(row['address']),
        'state': str(row['state']),
        'pincode': to_pin(row.get('pincode', '')),
        'match_status': 'UNMATCHED',
    }
    inst['city'] = extract_city_from_address(inst['address_sm'], inst['institute_name'])
    city = inst['city']
    if city and city.lower() not in inst['institute_name'].lower():
        inst['display_name'] = f"{inst['institute_name']}, {city}"
    else:
        inst['display_name'] = inst['institute_name']

    for prof_col in ['university', 'fee_yr1', 'fee_yr2', 'fee_yr3',
                     'stipend_yr1', 'stipend_yr2', 'stipend_yr3',
                     'hostel_male', 'hostel_female', 'bond_forfeit',
                     'pwbd_friendly', 'website', 'annual_fee']:
        inst[prof_col] = None

    confirmed.append(inst)

confirmed_df = pd.DataFrame(confirmed)
print(f"\nTotal institutes: {len(confirmed_df)}")
print(f"  EXACT: {len(confirmed_df[confirmed_df.match_status == 'EXACT'])}")
print(f"  OK_REVIEW: {len(confirmed_df[confirmed_df.match_status == 'OK_REVIEW'])}")
print(f"  UNMATCHED: {len(confirmed_df[confirmed_df.match_status == 'UNMATCHED'])}")

# Save intermediate CSV
confirmed_df.to_csv('/tmp/confirmed_institutes.csv', index=False)
print("Saved /tmp/confirmed_institutes.csv")


# ══════════════════════════════════════════════════════════════════════════════
# STEP 2: Create institutes table in DB
# ══════════════════════════════════════════════════════════════════════════════
print("\n" + "=" * 70)
print("STEP 2: Creating institutes table in DB")
print("=" * 70)

conn = sqlite3.connect(DB_PATH)
cur = conn.cursor()

# Drop and recreate
cur.execute("DROP TABLE IF EXISTS institute_mapping")
cur.execute("DROP TABLE IF EXISTS institutes")

cur.execute("""
    CREATE TABLE institutes (
        institute_code  INTEGER PRIMARY KEY,
        institute_name  TEXT NOT NULL,
        display_name    TEXT NOT NULL,
        address         TEXT,
        state           TEXT,
        pincode         TEXT,
        university      TEXT,
        fee_yr1         REAL,
        fee_yr2         REAL,
        fee_yr3         REAL,
        annual_fee      TEXT,
        stipend_yr1     TEXT,
        stipend_yr2     TEXT,
        stipend_yr3     TEXT,
        hostel_male     TEXT,
        hostel_female   TEXT,
        bond_forfeit    TEXT,
        pwbd_friendly   TEXT,
        website         TEXT,
        match_status    TEXT
    )
""")

cur.execute("CREATE INDEX ix_institutes_state ON institutes(state)")
cur.execute("CREATE INDEX ix_institutes_match_status ON institutes(match_status)")
cur.execute("CREATE INDEX ix_institutes_name ON institutes(institute_name)")

inserted = 0
for _, row in confirmed_df.iterrows():
    fee_yr1 = None
    if row.get('fee_yr1') is not None and not pd.isna(row.get('fee_yr1', None)):
        try: fee_yr1 = float(row['fee_yr1'])
        except: fee_yr1 = None

    fee_yr2 = None
    if row.get('fee_yr2') is not None and not pd.isna(row.get('fee_yr2', None)):
        try: fee_yr2 = float(row['fee_yr2'])
        except: fee_yr2 = None

    fee_yr3 = None
    if row.get('fee_yr3') is not None and not pd.isna(row.get('fee_yr3', None)):
        try: fee_yr3 = float(row['fee_yr3'])
        except: fee_yr3 = None

    def safe_str(v):
        if v is None or (isinstance(v, float) and math.isnan(v)):
            return None
        return str(v) if v else None

    cur.execute("""
        INSERT OR REPLACE INTO institutes VALUES (?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?)
    """, (
        int(row['institute_code']),
        str(row['institute_name']),
        str(row['display_name']),
        safe_str(row.get('address_sm')),
        safe_str(row.get('state')),
        safe_str(row.get('pincode')),
        safe_str(row.get('university')),
        fee_yr1, fee_yr2, fee_yr3,
        safe_str(row.get('annual_fee')),
        safe_str(row.get('stipend_yr1')),
        safe_str(row.get('stipend_yr2')),
        safe_str(row.get('stipend_yr3')),
        safe_str(row.get('hostel_male')),
        safe_str(row.get('hostel_female')),
        safe_str(row.get('bond_forfeit')),
        safe_str(row.get('pwbd_friendly')),
        safe_str(row.get('website')),
        str(row['match_status']),
    ))
    inserted += 1

conn.commit()
print(f"Inserted {inserted} institutes into DB")

# Verify
count = cur.execute("SELECT COUNT(*) FROM institutes").fetchone()[0]
print(f"DB institutes count: {count}")
for status, cnt in cur.execute("SELECT match_status, COUNT(*) FROM institutes GROUP BY match_status").fetchall():
    print(f"  {status}: {cnt}")


# ══════════════════════════════════════════════════════════════════════════════
# STEP 3: Match DB allotment institute_names to institutes
# ══════════════════════════════════════════════════════════════════════════════
print("\n" + "=" * 70)
print("STEP 3: Matching DB allotment institutes to institute list")
print("=" * 70)

# Get distinct institute names from allotments
db_institutes = [r[0] for r in cur.execute(
    "SELECT DISTINCT institute_name FROM allotments WHERE institute_name IS NOT NULL"
).fetchall()]
print(f"Distinct allotment institute names: {len(db_institutes)}")

# Build lookup structures for institutes table
inst_names = confirmed_df['institute_name'].tolist()
inst_names_norm = [norm(n) for n in inst_names]
inst_code_map = dict(zip(confirmed_df['institute_name'], confirmed_df['institute_code']))

# Also build a lookup by pincode
inst_pin_map = {}
for _, row in confirmed_df.iterrows():
    p = str(row.get('pincode', '')).strip()
    if p:
        inst_pin_map.setdefault(p, []).append(row['institute_name'])

# Create mapping table
cur.execute("""
    CREATE TABLE IF NOT EXISTS institute_mapping (
        db_institute_name   TEXT PRIMARY KEY,
        institute_code      INTEGER REFERENCES institutes(institute_code),
        match_confidence    TEXT,
        match_score         INTEGER
    )
""")
cur.execute("DELETE FROM institute_mapping")

# Get pincode and state for each DB institute
db_inst_info = {}
for name in db_institutes:
    info = cur.execute("""
        SELECT institute_name, state,
               COALESCE(institute_pincode, '') as pincode
        FROM allotments
        WHERE institute_name = ?
        LIMIT 1
    """, (name,)).fetchone()
    if info:
        db_inst_info[name] = {'state': info[1] or '', 'pincode': info[2] or ''}

# Match each DB institute to our list
mappings = []
for db_name in db_institutes:
    db_norm = norm(db_name)
    db_core = strip_generic(db_name)
    db_info = db_inst_info.get(db_name, {})
    db_state = db_info.get('state', '').lower().strip()
    db_pin = db_info.get('pincode', '').strip()

    # Try exact name match first
    exact_match = None
    for inst_name in inst_names:
        if norm(inst_name) == db_norm:
            exact_match = inst_name
            break

    if exact_match:
        code = inst_code_map[exact_match]
        mappings.append((db_name, code, 'EXACT', 100))
        continue

    # Fuzzy match
    results = process.extract(db_norm, inst_names_norm, scorer=fuzz.token_sort_ratio, limit=5)

    best_match = None
    best_score = 0
    best_core = 0
    best_confidence = 'UNKNOWN'

    for matched_str, score, idx in results:
        if score < 50:
            continue

        inst_name = inst_names[idx]
        inst_row = confirmed_df[confirmed_df['institute_name'] == inst_name].iloc[0]
        inst_state = str(inst_row.get('state', '')).lower().strip()
        inst_pin = str(inst_row.get('pincode', '')).strip()

        # Core score (generic words stripped)
        core_score = fuzz.token_set_ratio(db_core, strip_generic(inst_name))

        # Composite score
        composite = core_score
        if db_state and inst_state and db_state == inst_state:
            composite += 10
        if db_pin and inst_pin and db_pin == inst_pin:
            composite += 15
        if db_state and inst_state and db_state != inst_state:
            composite -= 20
        if db_pin and inst_pin and db_pin != inst_pin:
            composite -= 10

        if composite > best_score:
            best_score = composite
            best_core = core_score
            best_match = inst_name

            if core_score >= 90:
                best_confidence = 'EXACT'
            elif core_score >= 85 and composite >= 90:
                best_confidence = 'FUZZY'
            else:
                best_confidence = 'UNKNOWN'

    if best_match and best_confidence != 'UNKNOWN':
        code = inst_code_map[best_match]
        mappings.append((db_name, code, best_confidence, best_score))
    else:
        mappings.append((db_name, None, 'UNKNOWN', best_score))

# Insert mappings
for db_name, code, confidence, score in mappings:
    cur.execute(
        "INSERT OR REPLACE INTO institute_mapping VALUES (?, ?, ?, ?)",
        (db_name, code, confidence, score)
    )

conn.commit()

# Stats
total = len(mappings)
exact_count = sum(1 for _, _, c, _ in mappings if c == 'EXACT')
fuzzy_count = sum(1 for _, _, c, _ in mappings if c == 'FUZZY')
unknown_count = sum(1 for _, _, c, _ in mappings if c == 'UNKNOWN')

print(f"\nMapping results:")
print(f"  EXACT: {exact_count}")
print(f"  FUZZY: {fuzzy_count}")
print(f"  UNKNOWN: {unknown_count}")
print(f"  Total: {total}")


# ══════════════════════════════════════════════════════════════════════════════
# STEP 4: Export unknown institutes to Excel
# ══════════════════════════════════════════════════════════════════════════════
print("\n" + "=" * 70)
print("STEP 4: Exporting unknown institutes to Excel")
print("=" * 70)

unknown = [(db_name, score) for db_name, _, conf, score in mappings if conf == 'UNKNOWN']

if unknown:
    FONT_NAME = 'Arial'
    HDR_FILL = PatternFill('solid', start_color='1F497D')
    HDR_FONT = Font(name=FONT_NAME, bold=True, color='FFFFFF', size=10)
    WRAP_AL = Alignment(wrap_text=True, vertical='top')
    TOP_AL = Alignment(vertical='top')
    CENTER_AL = Alignment(horizontal='center', vertical='top')
    thin = Side(style='thin', color='BFBFBF')
    BORDER = Border(left=thin, right=thin, top=thin, bottom=thin)

    wb_out = openpyxl.Workbook()
    ws = wb_out.active
    ws.title = '不明大学リスト'

    headers = ['No.', 'DB大学名', 'DB州', 'DB PINCODE', '最良スコア', '備考']
    widths = [6, 50, 20, 12, 10, 30]

    for ci, (h, w) in enumerate(zip(headers, widths), 1):
        cell = ws.cell(1, ci, h)
        cell.font = HDR_FONT
        cell.fill = HDR_FILL
        cell.alignment = CENTER_AL
        cell.border = BORDER
        from openpyxl.utils import get_column_letter
        ws.column_dimensions[get_column_letter(ci)].width = w

    for ri, (db_name, score) in enumerate(sorted(unknown, key=lambda x: x[0]), 2):
        info = db_inst_info.get(db_name, {})
        vals = [ri - 1, db_name, info.get('state', ''), info.get('pincode', ''), score, '']
        for ci, val in enumerate(vals, 1):
            cell = ws.cell(ri, ci, val)
            cell.font = Font(name=FONT_NAME, size=9)
            cell.border = BORDER
            cell.alignment = TOP_AL

    ws.freeze_panes = 'A2'
    ws.auto_filter.ref = f"A1:F{ws.max_row}"

    out_path = os.path.join(OUTPUT_DIR, 'unknown_institutes.xlsx')
    wb_out.save(out_path)
    print(f"Saved {len(unknown)} unknown institutes to: {out_path}")
else:
    print("No unknown institutes found!")

conn.close()
print("\n✓ Done!")
